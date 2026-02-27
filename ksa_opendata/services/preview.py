"""Resource preview helpers with safety caps."""

from __future__ import annotations

import csv
import io
import json
from typing import Any, Dict

import httpx
from openpyxl import load_workbook

MAX_BYTES = 2_000_000


async def fetch_bytes(
    url: str,
    max_bytes: int = MAX_BYTES,
    timeout: float = 30.0,
) -> tuple[bytes, dict[str, str]]:
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        response = await client.get(url, headers={"User-Agent": "ksa-opendata-mcp/0.1"})
        response.raise_for_status()
        content = response.content[:max_bytes]
        headers = {k.lower(): v for k, v in response.headers.items()}
        return content, headers


def preview_csv(content: bytes, rows: int = 20) -> Dict[str, Any]:
    text = content.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    columns = reader.fieldnames or []
    out: list[Dict[str, Any]] = []
    for i, row in enumerate(reader):
        if i >= rows:
            break
        out.append(row)
    return {"type": "csv", "columns": columns, "rows": out, "returned_rows": len(out)}


def preview_json(content: bytes, rows: int = 20) -> Dict[str, Any]:
    text = content.decode("utf-8", errors="replace")
    payload = json.loads(text)
    if isinstance(payload, list):
        return {
            "type": "json",
            "shape": "list",
            "rows": payload[:rows],
            "returned_rows": min(len(payload), rows),
        }
    if isinstance(payload, dict):
        keys = list(payload.keys())[:100]
        return {"type": "json", "shape": "object", "keys": keys}
    return {"type": "json", "shape": type(payload).__name__}


def preview_xlsx(content: bytes, rows: int = 20) -> Dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return {"type": "xlsx", "columns": [], "rows": [], "returned_rows": 0}
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator, None)
    if not header:
        return {"type": "xlsx", "columns": [], "rows": [], "returned_rows": 0}
    columns = [str(c) if c is not None else "" for c in header]
    rows_out: list[Dict[str, Any]] = []
    for i, row in enumerate(iterator):
        if i >= rows:
            break
        rows_out.append({columns[j]: row[j] for j in range(min(len(columns), len(row)))})
    return {"type": "xlsx", "columns": columns, "rows": rows_out, "returned_rows": len(rows_out)}
